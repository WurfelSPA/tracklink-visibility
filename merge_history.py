"""merge_history.py — Fusiona reporte nuevo con historial (60 días). Visibility."""
import openpyxl, os, shutil, glob as glob_mod
from datetime import datetime, timedelta

LATEST   = 'latest.xlsx'
HIST     = 'REPORTE DE RECORRIDO.xlsx'
MAX_DAYS = 60

def get_header_row(ws, max_search=10):
    for r in range(1, max_search + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(str(v).strip() == 'Alias' for v in row_vals if v):
            col = {str(v).strip(): i+1 for i,v in enumerate(row_vals) if v}
            return r, col
    return None, {}

def extract_rows(ws, header_row, col):
    rows = []
    alias_col = col.get('Alias', 1)
    for r in range(header_row + 1, ws.max_row + 1):
        if not ws.cell(r, alias_col).value:
            continue
        rows.append({h: ws.cell(r, col[h]).value for h in col})
    return rows

def row_key(row):
    alias  = str(row.get('Alias','') or '').strip()
    estado = str(row.get('Estado','') or '').strip()
    seq    = str(row.get('Secuencial','') or '').strip()
    return (alias, estado, seq)

def parse_fecha(val):
    if val is None: return None
    if isinstance(val, datetime): return val
    for fmt in ('%Y/%m/%d %H:%M:%S','%Y-%m-%d %H:%M:%S','%d/%m/%Y %H:%M:%S'):
        try: return datetime.strptime(str(val), fmt)
        except: pass
    return None

def main():
    bulk   = sorted(glob_mod.glob('bulk_*.xlsx'))
    sources = bulk or ([LATEST] if os.path.exists(LATEST) else [])
    if not sources:
        print('ERROR: sin archivos de entrada'); exit(1)

    new_rows = []
    for src in sources:
        try:
            wb = openpyxl.load_workbook(src)
            if 'Detalle 1' not in wb.sheetnames:
                print(f'  [SKIP] {src}: sin Detalle 1'); continue
            ws = wb['Detalle 1']
            hr, col = get_header_row(ws)
            if not hr: print(f'  [SKIP] {src}: sin encabezado'); continue
            rows = extract_rows(ws, hr, col)
            new_rows.extend(rows)
            print(f'  {src}: {len(rows)} filas')
        except Exception as e:
            print(f'  [WARN] {src}: {e}')
    print(f'Total filas nuevas: {len(new_rows)}')

    existing_rows, existing_keys = [], set()
    if os.path.exists(HIST):
        try:
            wb_old = openpyxl.load_workbook(HIST)
            if 'Detalle 1' in wb_old.sheetnames:
                ws_old = wb_old['Detalle 1']
                hr_old, col_old = get_header_row(ws_old)
                if hr_old:
                    existing_rows = extract_rows(ws_old, hr_old, col_old)
                    existing_keys = {row_key(r) for r in existing_rows}
                    print(f'Historial: {len(existing_rows)} filas')
        except Exception as e:
            print(f'[WARN] historial: {e}')

    added = 0
    for row in new_rows:
        k = row_key(row)
        if k not in existing_keys:
            existing_rows.append(row); existing_keys.add(k); added += 1
    print(f'Nuevas: {added}')

    existing_rows.sort(key=lambda r: parse_fecha(r.get('Fecha de Inicio')) or datetime.min)

    if existing_rows:
        last = parse_fecha(existing_rows[-1].get('Fecha de Inicio'))
        if last:
            cutoff = last.replace(hour=0,minute=0,second=0) - timedelta(days=MAX_DAYS)
            before = len(existing_rows)
            existing_rows = [r for r in existing_rows
                             if (parse_fecha(r.get('Fecha de Inicio')) or datetime.min) >= cutoff]
            print(f'Recorte: {before} → {len(existing_rows)} filas')

    template = sources[0]
    shutil.copy(template, HIST)
    wb_out = openpyxl.load_workbook(HIST)
    if 'Detalle 1' not in wb_out.sheetnames:
        print('ERROR: sin Detalle 1 en plantilla'); exit(1)
    ws_out = wb_out['Detalle 1']
    hr_out, col_out = get_header_row(ws_out)
    ws_out.delete_rows(hr_out + 1, ws_out.max_row - hr_out)
    for i, row in enumerate(existing_rows):
        for h, cidx in col_out.items():
            ws_out.cell(hr_out + 1 + i, cidx).value = row.get(h)
    wb_out.save(HIST)
    print(f'✅ {HIST}: {len(existing_rows)} filas totales')

if __name__ == '__main__':
    main()
